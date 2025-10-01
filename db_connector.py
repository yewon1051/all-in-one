import psycopg2
import csv
import json
from datetime import datetime
import os

class PostgreSQLConnector:
    def __init__(self):
        self.connection = None
        self.cursor = None
        
        # 데이터베이스 연결 정보
        self.db_config = {
            'host': 'butfitseoul-replica.cjilul7too7t.ap-northeast-2.rds.amazonaws.com',
            'port': 5432,
            'database': 'master_20221217',
            'user': 'ywlee',
            'password': 'Tha1xia5Poo0aethei0eifauz8udo4oh'
        }
    
    def connect(self):
        """데이터베이스에 연결"""
        try:
            self.connection = psycopg2.connect(**self.db_config)
            self.cursor = self.connection.cursor()
            print("✅ PostgreSQL 데이터베이스에 성공적으로 연결되었습니다!")
            return True
        except psycopg2.Error as e:
            print(f"❌ 데이터베이스 연결 실패: {e}")
            return False
    
    def disconnect(self):
        """데이터베이스 연결 해제"""
        if self.cursor:
            self.cursor.close()
        if self.connection:
            self.connection.close()
        print("🔌 데이터베이스 연결이 해제되었습니다.")
    
    def execute_query(self, query):
        """쿼리 실행하고 결과 반환"""
        if not self.connection:
            print("❌ 데이터베이스에 연결되지 않았습니다. connect()를 먼저 호출하세요.")
            return None
        
        try:
            self.cursor.execute(query)
            
            # SELECT 쿼리인 경우 결과 반환
            if query.strip().upper().startswith('SELECT'):
                columns = [desc[0] for desc in self.cursor.description]
                rows = self.cursor.fetchall()
                return {
                    'columns': columns,
                    'data': rows,
                    'row_count': len(rows)
                }
            else:
                # INSERT, UPDATE, DELETE 등의 경우
                self.connection.commit()
                return {
                    'message': '쿼리가 성공적으로 실행되었습니다.',
                    'affected_rows': self.cursor.rowcount
                }
                
        except psycopg2.Error as e:
            print(f"❌ 쿼리 실행 실패: {e}")
            return None
    
    def save_to_csv(self, query_result, filename=None):
        """쿼리 결과를 CSV 파일로 저장"""
        if not query_result or 'columns' not in query_result:
            print("❌ 저장할 데이터가 없습니다.")
            return False
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"query_result_{timestamp}.csv"
        
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # 헤더 작성
                writer.writerow(query_result['columns'])
                
                # 데이터 작성
                writer.writerows(query_result['data'])
            
            print(f"✅ CSV 파일이 저장되었습니다: {filename}")
            print(f"📊 총 {query_result['row_count']}개의 행이 저장되었습니다.")
            return True
            
        except Exception as e:
            print(f"❌ CSV 저장 실패: {e}")
            return False
    
    def save_to_excel(self, query_result, filename=None):
        """쿼리 결과를 Excel 파일로 저장"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            if not query_result or 'columns' not in query_result:
                print("❌ 저장할 데이터가 없습니다.")
                return False
            
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"query_result_{timestamp}.xlsx"
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Query Result"
            
            # 헤더 스타일 설정
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # 헤더 작성
            for col_num, column in enumerate(query_result['columns'], 1):
                cell = ws.cell(row=1, column=col_num, value=column)
                cell.font = header_font
                cell.fill = header_fill
            
            # 데이터 작성
            for row_num, row_data in enumerate(query_result['data'], 2):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            print(f"✅ Excel 파일이 저장되었습니다: {filename}")
            print(f"📊 총 {query_result['row_count']}개의 행이 저장되었습니다.")
            return True
            
        except ImportError:
            print("❌ openpyxl이 설치되지 않았습니다. 'pip install openpyxl'을 실행하세요.")
            return False
        except Exception as e:
            print(f"❌ Excel 저장 실패: {e}")
            return False
    
    def get_table_list(self):
        """데이터베이스의 테이블 목록 조회"""
        query = """
        SELECT table_name, table_type 
        FROM information_schema.tables 
        WHERE table_schema = 'public'
        ORDER BY table_name;
        """
        return self.execute_query(query)
    
    def get_table_info(self, table_name):
        """특정 테이블의 컬럼 정보 조회"""
        query = f"""
        SELECT column_name, data_type, is_nullable, column_default
        FROM information_schema.columns
        WHERE table_name = '{table_name}'
        ORDER BY ordinal_position;
        """
        return self.execute_query(query)


